import os, json
import openpyxl
import instaloader
import shutil

try:
    from instaloader import ConnectionException, Instaloader, Profile, Post
except ModuleNotFoundError:
    raise SystemExit("Instaloader not found.\n  pip install [--user] instaloader")



class Scraper():

    def __init__(self):
        try:
            with open('config.txt') as f:
                lines = f.readlines()
            self.user=lines[0].strip()
            self.password=lines[1].strip()
        except OSError :
            return  ('There is no config.txt file. Create it in the current path.try again')  

    def Exportfilecomment(self,type,link):
        try:
            
            # name=link[-40:]
            # name=name[:11]
            json_files = [pos_json for pos_json in os.listdir('test/') if pos_json.endswith('.json')]
            f = open("test/"+json_files[0],'r') 
            data = json.load(f)
            list2=[] 
            for i in data:
                name=i['owner']['username']
                list2.append(name)
            book = openpyxl.Workbook()
            sheet = book.get_sheet_by_name('Sheet')
            i=2
            sheet.cell(row=1, column=1).value = 'username'
            for x in list2:
                sheet.cell(row=i, column=1).value = x
                i=i+1
            book.save('Sample.xlsx')
            return ('Sample.xlsx')
        except Exception as ex :
            return ex
    def Exportfilelike(self,type,link):
        try:        
            
            a=link[-40:]
            test1 = Instaloader(download_pictures=False,download_videos=False,download_video_thumbnails=False)
            test1.login(self.user,self.password)
            post = Post.from_shortcode(test1.context, a[0:11])
            book = openpyxl.Workbook()
            sheet = book.get_sheet_by_name('Sheet')
            i=2
            sheet.cell(row=1, column=1).value = 'username'
            for x in post.get_likes():
                sheet.cell(row=i, column=1).value = x.username
                i=i+1
            book.save('Samplelike.xlsx')
            return ('Samplelike.xlsx')
        except Exception as ex :
            return ex
    def Exportfilefollower(self,type,link):
        try:        
            a=link[-40:]
            test1 = Instaloader(download_pictures=False,download_videos=False,download_video_thumbnails=False)
            test1.login(self.user,self.password)
            post = Post.from_shortcode(test1.context, a[0:11])
            profile = instaloader.Profile.from_username(test1.context, post.owner_username)
            book = openpyxl.Workbook()
            sheet = book.get_sheet_by_name('Sheet')
            i=2
            sheet.cell(row=1, column=1).value = 'username'
            for x in profile.get_followers():
                sheet.cell(row=i, column=1).value = x.username
                i=i+1
            book.save('Samplef.xlsx')
            return ('Samplef.xlsx')
        except Exception :
             return  ('Login to instagram fail.Please wait a few minutes before you try again')  #render_template  ("login.html",src="",list=[False,False,False,False],post="",text="",count="",flag=False,flag1=False,flag2=False)
   
    def Exportfilemention(self,type,link):
        try:        

            # name=link[-40:]
            # name=name[:11]
            json_files = [pos_json for pos_json in os.listdir('test/') if pos_json.endswith('.json')]
            f = open("test/"+json_files[0],'r') 
            data = json.load(f)
            list2=[] 
            list1=[]
            for i in data:
                name=i['owner']['username']
                text=i['text']
                s1=""
                s=""
                i=0
                j=0
                while  i<len(text ):
                    s=""
                    if ord(text[i])==64:
                        s +=text[i]
                        i=i+1
                        while i<len(text):
                            if ord(text[i])!=32:
                                s +=text[i]
                                i=i+1
                            else:
                                break
                                
                        s1 +=','+s
                        j=j+1

                    i=i+1
                if j>0:
                    list2.append(name)
                    list2.append(s1)
                    list2.append(j)
                    list1.append(list2)
                list2=[]
            book = openpyxl.Workbook()
            sheet = book.get_sheet_by_name('Sheet')
            i=2
            j=1
            sheet.cell(row=1, column=1).value = 'Username'
            sheet.cell(row=1, column=2).value = 'Mention'
            sheet.cell(row=1, column=3).value = 'Count'
            for x in list1:
                sheet.cell(row=i, column=j).value = x[0]
                j=j+1
                sheet.cell(row=i, column=j).value = x[1]
                j=j+1
                sheet.cell(row=i, column=j).value = x[2]
                j=1
                i=i+1
            book.save('Samplemention.xlsx')
            return ('Samplemention.xlsx')
        except Exception :
            return  ('Login to instagram fail.Please wait a few minutes before you try again')  #render_template  ("login.html",src="",list=[False,False,False,False],post="",text="",count="",flag=False,flag1=False,flag2=False)
   

    def scraperpost(self, shortpost):
        try:

            test1 = Instaloader(download_pictures=False,download_videos=False,download_video_thumbnails=False)
            test1.login(self.user,self.password)
            post = Post.from_shortcode(test1.context, shortpost)
            test1.download_post(post, target='test')
            return (post.owner_username)
            
        except Exception as ex :
            return ex

    def textcomment(self):
        try:
            path_to_txt = 'test/'
            txt_files = [pos_json for pos_json in os.listdir(path_to_txt) if pos_json.endswith('.txt')]
            f = open("test/"+str(txt_files[len(txt_files)-1]),'r') 
            content=f.read()
            return content

        except: 
            return False
        
    def countcomment(self,post):
        try:
            json_files = [pos_json for pos_json in os.listdir('test/') if pos_json.endswith('.json')]
            f = open("test/"+json_files[0],'r') 
            data = json.load(f) 
            return (len(data))

        except: 
            return False

    def countlike(self,post):
        try:
            a=post[-40:]
            test1 = Instaloader(download_pictures=False,download_videos=False,download_video_thumbnails=False)
            test1.login(self.user,self.password)
            post = Post.from_shortcode(test1.context, a[0:11])
            return (len(set(post.get_likes())))

        except: 
            return False

    def countf(self,shortpost):
        try:

            test1 = Instaloader(download_pictures=False,download_videos=False,download_video_thumbnails=False)
            test1.login(self.user,self.password)
            profile = instaloader.Profile.from_username(test1.context, shortpost)
            return (len(set(profile.get_followers())))

        except: 
            return False
    def countmention(self,post):
        try:
            
            json_files = [pos_json for pos_json in os.listdir('test/') if pos_json.endswith('.json')]

            f = open("test/"+json_files[0],'r') 
            data = json.load(f) 
            countm=0
            for x in data:
                text=x['text']
                s1=""
                s=""
                i=0
                j=0
                while  i<len(text ):
                    s=""
                    if ord(text[i])==64:
                        s +=text[i]
                        i=i+1
                        while i<len(text):
                            if ord(text[i])!=32:
                                s +=text[i]
                                i=i+1
                            else:
                                break
                                
                        s1 +=','+s
                        j=j+1
                        # countm =countm+1

                    i=i+1
                if j>0:
                    countm =countm+1

                    
            return (countm)
        except: 
            return False
    def get_list_like(self,link):
        try:
            a=link[-40:]
            listt=[]
            test1 = Instaloader(download_pictures=False,download_videos=False,download_video_thumbnails=False)
            test1.login(self.user,self.password)
            post = Post.from_shortcode(test1.context, a[0:11])
            for x in list(post.get_likes()):
                 listt.append(x.username)
            return (listt)

        except Exception as ex: 
            return ex
    def get_list_follower(self,link):
        try:
            a=link[-40:]
            listt=[]
            test1 = Instaloader(download_pictures=False,download_videos=False,download_video_thumbnails=False)
            test1.login(self.user,self.password)
            post = Post.from_shortcode(test1.context, a[0:11])
            profile = instaloader.Profile.from_username(test1.context, post.owner_username)
            for x in list(profile.get_followers()):
                 listt.append(x.username)
            return (listt)

        except Exception as ex: 
            return ex
    def get_list_comment(self,link):
        try:
            
            
            list1=[]
            json_files = [pos_json for pos_json in os.listdir('test/') if pos_json.endswith('.json')]
            f = open("test/"+json_files[0],'r') 
            data = json.load(f) 
            for x in data:
                list1.append(x['owner']['username'])
            return (list1)
        except Exception as ex: 
            return ex
    def get_list_mention(self,link):
        try:
            
            json_files = [pos_json for pos_json in os.listdir('test/') if pos_json.endswith('.json')]
            f = open("test/"+json_files[0],'r') 
            data = json.load(f)
            list2=[] 
            list1=[]
            for i in data:
                name=i['owner']['username']
                text=i['text']
                s1=""
                s=""
                i=0
                j=0
                while  i<len(text ):
                    s=""
                    if ord(text[i])==64:
                        s +=text[i]
                        i=i+1
                        while i<len(text):
                            if ord(text[i])!=32:
                                s +=text[i]
                                i=i+1
                            else:
                                break
                                
                        s1 +=','+s
                        j=j+1

                    i=i+1
                if j>0:
                    list1.append(name)
                if j>=5 :
                    c=j/5
                    i=1
                    for i in i<=c:
                        list1.append(name)

                   
                
            return(list1)
        except Exception as ex: 
            return ex

